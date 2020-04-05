import discord
import requests, xmltodict, json
from datetime import datetime, timedelta
import pandas as pd
import xlrd
import openpyxl
import random


client = discord.Client()
year = datetime.now().strftime("%Y")
month = datetime.now().strftime("%m")
day = datetime.now().strftime("%d")
datered = year + month + day
dater = int(datered)
basetimeget = 0
skybasetimeget = 0
dateget = 0
daymonths = datetime.now().month
daymonth = str(daymonths)
daydays = datetime.now().day
dayday = str(daydays)
dayweekday = datetime.now().weekday()
dayweekdayget = ""
daymonthget = daymonth + "월"
daydayget = dayday + "일"
hourer = datetime.now().strftime("%H")
minuter = datetime.now().strftime("%M")
timer = hourer + minuter
timerget = int(timer)
PTY = 0
REH = 0
RN1 = 0
T1H = 0
placex = 55
placey = 124
PTYd = int(PTY)
placename = '인천광역시'
key = "PDWjrGNh4fjziywH%2Fnh6TDZc9SWkk04ADa4c%2BcQtEc0R9wCHSTjGRxQ5qUZbpOFQuVL2Cd9bNwy0IqCBPl7yNw%3D%3D"
key2 = 'RGAPI-2a7a8f27-2022-40b6-afbe-135198b1553f'
placework = xlrd.open_workbook('weather.xlsx')
placesheet = placework.sheet_by_name('sheet1')


@client.event
async def on_ready():
    await client.change_presence(activity=discord.Game(name=' 강연 ㅣ %명령어 '))
    print("ready")


@client.event
async def on_message(message):
    if message.content.startswith("%용안"):
        await message.channel.send(file=discord.File('세종.jpg'))
        await message.channel.send('나의 얼굴이 어떤가?')
    if message.content.startswith("%주사위"):
        randomsd = message.content.split(" ")[1]
        rint = random.randint(1, int(randomsd))
        rinter = str(rint)
        username = message.author.name
        await message.channel.send('**' + username + '**님의 주사위에서는 **__' + rinter + '__**가 나왔습니다!!')
    if message.content.startswith("%명령어"):
        combed = discord.Embed(colour=0x00FFFF, title='세종 이도의 명령들')
        combed.add_field(name='%명령어', value='현재 사용 가능한 명령어들을 확인합니다')
        combed.add_field(name='%용안', value='세종대왕님의 용안을 뵙는 명령어입니다')
        combed.add_field(name='%날씨 - 제작중', value='인천광역시의 날씨를 확인합니다')
        combed.add_field(name='%주사위 [숫자]', value='1부터 [숫자]사이의 랜덤 숫자를 골라줍니다')
        combed.add_field(name='업데이트 예정 내용', value='1. 날씨 서비스 지역별 제공 - 이모지를 이용한 지역 선정 \n2. 롤 정보 조회 서비스 - 닉네임 검색 우선', inline=False)
        combed.set_footer(text='현재 개발 중인 봇입니다. Resource produced by Data.go.kr / Riotgames.com')
        await message.channel.send(embed=combed)
    if message.content.startswith("%채널"):
        channel = message.content[4:22]
        msg = message.content[23:]
        await client.get_channel(int(channel)).send(msg)
    if message.content.startswith("%테스트"):
        await message.channel.send('ss')
    if message.content.startswith('%롤유저'):
        def champer():
            usernick = message.content.split(" ")[1]
            urls = "https://kr.api.riotgames.com/lol/summoner/v4/summoners/by-name/{}?api_key={}".format(usernick, key2)

            contents = requests.get(urls).content
            dicts = json.dumps(contents, ensure_ascii=False)
            print(dicts)
        champer()
    if message.content.startswith("%날씨"):
        def weather():
            if 0 <= timerget <= 2400:
                if timerget <= 40:
                    basetimeget = '2300'
                    dateget = dater - 1
                elif 40 < timerget <= 140:
                    basetimeget = '0000'
                    dateget = dater
                elif 140 < timerget <= 240:
                    basetimeget = '0100'
                    dateget = dater
                elif 240 < timerget <= 340:
                    basetimeget = '0200'
                    dateget = dater
                elif 340 < timerget <= 440:
                    basetimeget = '0300'
                    dateget = dater
                elif 440 < timerget <= 540:
                    basetimeget = '0400'
                    dateget = dater
                elif 540 < timerget <= 640:
                    basetimeget = '0500'
                    dateget = dater
                elif 640 < timerget <= 740:
                    basetimeget = '0600'
                    dateget = dater
                elif 740 < timerget <= 840:
                    basetimeget = '0700'
                    dateget = dater
                elif 840 < timerget <= 940:
                    basetimeget = '0800'
                    dateget = dater
                elif 940 < timerget <= 1040:
                    basetimeget = '0900'
                    dateget = dater
                elif 1040 < timerget <= 1140:
                    basetimeget = '1000'
                    dateget = dater
                elif 1140 < timerget <= 1240:
                    basetimeget = '1100'
                    dateget = dater
                elif 1240 < timerget <= 1340:
                    basetimeget = '1200'
                    dateget = dater
                elif 1340 < timerget <= 1440:
                    basetimeget = '1300'
                    dateget = dater
                elif 1440 < timerget <= 1540:
                    basetimeget = '1400'
                    dateget = dater
                elif 1540 < timerget <= 1640:
                    basetimeget = '1500'
                    dateget = dater
                elif 1640 < timerget <= 1740:
                    basetimeget = '1600'
                    dateget = dater
                elif 1740 < timerget <= 1840:
                    basetimeget = '1700'
                    dateget = dater
                elif 1840 < timerget <= 1940:
                    basetimeget = '1800'
                    dateget = dater
                elif 1940 < timerget <= 2040:
                    basetimeget = '1900'
                    dateget = dater
                elif 2040 < timerget <= 2140:
                    basetimeget = '2000'
                    dateget = dater
                elif 2140 < timerget <= 2240:
                    basetimeget = '2100'
                    dateget = dater
                elif 2240 < timerget <= 2340:
                    basetimeget = '2200'
                    dateget = dater
                elif timerget > 2340:
                    basetimeget = '2300'
                    dateget = dater


            url = "http://apis.data.go.kr/1360000/VilageFcstInfoService/getUltraSrtNcst?serviceKey={}&pageNo=1&startPage=1&numOfRows=10&pageSize=10&dataType=XML&base_date={}&base_time={}&nx={}&ny={}".format(
                key, dateget, basetimeget, placex, placey)

            content = requests.get(url).content
            dict = xmltodict.parse(content)
            jsonstring = json.dumps(dict['response']['body'], ensure_ascii=False)
            json0bj = json.loads(jsonstring)

            for item in json0bj['items']['item']:
                print(item)

            df1 = pd.DataFrame(json0bj['items']['item'])
            file = open("./weather.json", "w+")
            file.write(json.dumps(json0bj['items']['item']))

            def save1(df1, weather):
                writer = pd.ExcelWriter(weather)
                df1.to_excel(writer, "sheet1")
                writer.save()

            save1(df1, "weather.xlsx")
        weather()
        workbook = xlrd.open_workbook('weather.xlsx')
        worksheet = workbook.sheet_by_name('sheet1')
        list = worksheet._cell_values
        for row in list[1:2]:
            PTY2 = row[6]
        for row in list[2:3]:
            REH2 = row[6]
        for row in list[3:4]:
            RN12 = row[6]
        for row in list[4:5]:
            T1H2 = row[6]
        PTY = PTY2
        REH = REH2
        RN1 = RN12
        T1H = T1H2


        print(PTY, REH, RN1, T1H)

        def weathers():

            if 0 <= timerget <= 2400:
                if timerget <= 210:
                    skybasetimeget = '2300'
                    dateget = dater - 1
                elif 210 < timerget <= 510:
                    skybasetimeget = '0200'
                    dateget = dater
                elif 510 < timerget <= 810:
                    skybasetimeget = '0500'
                    dateget = dater
                elif 810 < timerget <= 1110:
                    skybasetimeget = '0800'
                    dateget = dater
                elif 1110 < timerget <= 1410:
                    skybasetimeget = '1100'
                    dateget = dater
                elif 1410 < timerget <= 1710:
                    skybasetimeget = '1400'
                    dateget = dater
                elif 1710 < timerget <= 2010:
                    skybasetimeget = '1700'
                    dateget = dater
                elif 2010 < timerget <= 2310:
                    skybasetimeget = '2000'
                    dateget = dater
                elif timerget > 2310:
                    skybasetimeget = '2300'
                    dateget = dater
            urls = "http://apis.data.go.kr/1360000/VilageFcstInfoService/getVilageFcst?serviceKey={}&pageNo=1&startPage=1&numOfRows=10&pageSize=10&dataType=XML&base_date={}&base_time={}&nx={}&ny={}".format(
                key, dateget, skybasetimeget, placex, placey)

            contents = requests.get(urls).content
            dicts = xmltodict.parse(contents)
            jsonstrings = json.dumps(dicts['response']['body'], ensure_ascii=False)
            json0bjs = json.loads(jsonstrings)

            for item in json0bjs['items']['item']:
                print(item)

            df1s = pd.DataFrame(json0bjs['items']['item'])
            files = open("./weathers.json", "w+")
            files.write(json.dumps(json0bjs['items']['item']))

            def save1s(df1s, weathers):
                writers = pd.ExcelWriter(weathers)
                df1s.to_excel(writers, "sheet1")
                writers.save()

            save1s(df1s, "weathers.xlsx")
        weathers()

        workbooks = xlrd.open_workbook('weathers.xlsx')
        worksheets = workbooks.sheet_by_name('sheet1')

        D5 = worksheets.cell(rowx=4, colx=3).value

        D7 = worksheets.cell(rowx=6, colx=3).value

        if D5 == 'SKY':
            SKY = worksheets.cell(rowx=4, colx=6).value
        elif D7 == 'SKY':
            SKY = worksheets.cell(rowx=6, colx=6).value

        POP = worksheets.cell(rowx=1, colx=6).value

        if SKY == '1':
            colors = 0xFFFF37
            weatherpic = 'sun.png'
        elif SKY == '3':
            colors = 0x413288
            weatherpic = 'cloud.png'
        elif SKY == '4':
            colors = 0x808080
            if PTYd == 0:
                weatherpic = 'cloud.png'
            elif PTYd == 1:
                weatherpic = 'rain.png'
            elif PTYd == 2:
                weatherpic = 'swain.png'
            elif PTYd == 3:
                weatherpic = 'snow.png'
            elif PTYd == 4:
                weatherpic = 'rain.png'

        embed = discord.Embed(colour=colors, title=placename + '의 현재 날씨는?')
        embed.add_field(name='기온', value=T1H + '℃', inline=True)
        if SKY == '1':
            embed.add_field(name='현재 날씨', value=' 맑음', inline=True)
        elif SKY == '3':
            embed.add_field(name='현재 날씨', value=' 구름 많음', inline=True)
        elif SKY == '4':
            embed.add_field(name='현재 날씨', value=' 흐림 \ 비', inline=True)
        embed.add_field(name='습도', value=REH + '%', inline=True)
        embed.add_field(name='강수 확률', value=POP + '%', inline=True)
        if PTYd > 0:
            embed.add_field(name='시간 당 강수량', value=RN1 + 'mm', inline=True)
            if PTYd == 1:
                embed.add_field(name='강수 형태', value='비', inline=True)
            if PTYd == 2:
                embed.add_field(name='강수 형태', value='진눈개비', inline=True)
            if PTYd == 3:
                embed.add_field(name='강수 형태', value='눈', inline=True)
            if PTYd == 4:
                embed.add_field(name='강수 형태', value='소나기', inline=True)
        embed.set_footer(text='다음 저작권은 기상청에 있습니다')
        await message.channel.send(embed=embed, file=discord.File(weatherpic))


client.run("NjkyNjU0MjY1MDEyOTc3Njk1.Xn3G3g.V48dLa9BErCz2cC6S8f7OKuwuxQ")
